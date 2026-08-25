r"""Build a file-level and variable-level report from CDE-ID reconciled workbooks.

The report separates three questions that should not be collapsed into one
"accuracy" percentage:

1. Was a final CDE candidate accepted by reconciliation?
2. How conceptually similar is the study variable to the final CDE?
3. How closely do the study and official CDE encodings align?

Example (PowerShell):

python .\build_cde_mapping_report.py `
  --input-dir "C:\path\to\out\v6-test\reconciled" `
  --run-report "C:\path\to\run_report_2026-08-19.csv" `
  --cde-json "C:\path\to\All_HEALPAINCDEsDD_JSON.json" `
  --config "C:\path\to\config_prestep.ini" `
  --output "C:\path\to\cde_mapping_metrics_report.xlsx"
"""

from __future__ import annotations

import argparse
import configparser
import json
import math
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from fuzzywuzzy import fuzz as fuzzy_fuzz
except ImportError:
    fuzzy_fuzz = None


DEFAULTS = {
    "var_name_weight": 0.25,
    "field_label_weight": 0.50,
    "form_context_weight": 0.15,
    "crf_context_weight": 0.10,
    "crf_concept_bonus": 8.0,
    "concept_high": 75.0,
    "concept_medium": 50.0,
    "fidelity_high": 80.0,
    "fidelity_medium": 50.0,
}

HEAL_PURPLE = "532565"
HEAL_MAGENTA = "982568"
HEAL_DARK = "373A3C"
HEAL_LIGHT = "F4EEF6"
WHITE = "FFFFFF"
LIGHT_GRAY = "E7E7E7"
GREEN = "D9EAD3"
YELLOW = "FFF2CC"
RED = "F4CCCC"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Summarize CDE-ID reconciled workbooks into one Excel report."
    )
    parser.add_argument(
        "--input-dir",
        required=True,
        type=Path,
        help="Folder containing *_cde_reconciled*.xlsx workbooks.",
    )
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Path for the generated .xlsx report.",
    )
    parser.add_argument(
        "--run-report",
        type=Path,
        default=None,
        help="Optional batch run report CSV to include and join by output filename.",
    )
    parser.add_argument(
        "--cde-json",
        type=Path,
        default=None,
        help="Optional official HEAL CDE JSON used to recompute post-recon scores.",
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Optional config_prestep.ini used for the notebook's weights and thresholds.",
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Search subfolders under --input-dir.",
    )
    return parser.parse_args()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip() == ""


def clean_text(value: Any) -> str:
    return "" if is_blank(value) else str(value).strip()


def value_to_text(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, (list, tuple, set)):
        return " | ".join(clean_text(item) for item in value if not is_blank(item))
    if isinstance(value, dict):
        return " | ".join(
            f"{clean_text(key)}={clean_text(item)}"
            for key, item in value.items()
            if not is_blank(item)
        )
    return clean_text(value)


def normalize_string(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-zA-Z0-9\s=]", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_identifier(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def normalize_crf_label(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    junk = {"questionnaire", "scale", "inventory", "form", "survey", "assessment"}
    return " ".join(token for token in text.split() if token not in junk)


def _ratio(left: str, right: str) -> int:
    return round(100 * SequenceMatcher(None, left, right).ratio())


def fallback_token_set_ratio(left: str, right: str) -> int:
    """Dependency-free approximation of fuzzywuzzy.fuzz.token_set_ratio."""
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    intersection = left_tokens & right_tokens
    left_only = left_tokens - intersection
    right_only = right_tokens - intersection

    shared = " ".join(sorted(intersection)).strip()
    left_combined = " ".join(sorted(intersection | left_only)).strip()
    right_combined = " ".join(sorted(intersection | right_only)).strip()
    pairs = [
        (shared, left_combined),
        (shared, right_combined),
        (left_combined, right_combined),
    ]
    scores = [_ratio(a, b) for a, b in pairs if a and b]
    return max(scores) if scores else 0


def similarity_score(left: Any, right: Any) -> int:
    left_norm = normalize_string(left)
    right_norm = normalize_string(right)
    if not left_norm or not right_norm:
        return 0
    if fuzzy_fuzz is not None:
        return int(fuzzy_fuzz.token_set_ratio(left_norm, right_norm))
    return fallback_token_set_ratio(left_norm, right_norm)


def safe_float(value: Any) -> float | None:
    if is_blank(value):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return None if math.isnan(result) else result


def first_existing(columns: list[str], candidates: list[str]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    return None


def read_settings(config_path: Path | None) -> dict[str, float]:
    settings = DEFAULTS.copy()
    if config_path is None:
        return settings
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")

    config = configparser.ConfigParser()
    config.read(config_path, encoding="utf-8")

    def get_float(section: str, option: str, default_key: str) -> float:
        return config.getfloat(section, option, fallback=settings[default_key])

    settings.update(
        {
            "var_name_weight": get_float("ConceptWeights", "var_name_weight", "var_name_weight"),
            "field_label_weight": get_float("ConceptWeights", "field_label_weight", "field_label_weight"),
            "form_context_weight": get_float("ConceptWeights", "form_context_weight", "form_context_weight"),
            "crf_context_weight": get_float("ConceptWeights", "crf_context_weight", "crf_context_weight"),
            "crf_concept_bonus": get_float("ConceptRetrieval", "crf_concept_bonus", "crf_concept_bonus"),
            "concept_high": get_float("ConceptThresholds", "high", "concept_high"),
            "concept_medium": get_float("ConceptThresholds", "medium", "concept_medium"),
            "fidelity_high": get_float("FidelityThresholds", "high", "fidelity_high"),
            "fidelity_medium": get_float("FidelityThresholds", "medium", "fidelity_medium"),
        }
    )
    return settings


def load_cde_lookup(json_path: Path | None) -> dict[str, list[dict[str, Any]]]:
    if json_path is None:
        return {}
    if not json_path.exists():
        raise FileNotFoundError(f"CDE JSON not found: {json_path}")

    with json_path.open("r", encoding="utf-8") as handle:
        raw = json.load(handle)

    if isinstance(raw, dict):
        entries = []
        for key, value in raw.items():
            if isinstance(value, dict):
                entry = dict(value)
                entry["_json_key"] = key
                entries.append(entry)
    elif isinstance(raw, list):
        entries = [dict(value) for value in raw if isinstance(value, dict)]
    else:
        raise ValueError("CDE JSON must contain an object or a list of objects.")

    lookup: dict[str, list[dict[str, Any]]] = {}
    for entry in entries:
        keys = [
            entry.get("_json_key"),
            entry.get("Variable Name"),
            entry.get("variable_name"),
            entry.get("CDE Name"),
            entry.get("cde_name"),
        ]
        normalized_keys = {normalize_identifier(key) for key in keys}
        for normalized_key in normalized_keys:
            if normalized_key:
                lookup.setdefault(normalized_key, []).append(entry)
    return lookup


def choose_official_entry(
    lookup: dict[str, list[dict[str, Any]]],
    selected_variable: Any,
    selected_cde_name: Any,
) -> tuple[dict[str, Any] | None, str]:
    if not lookup:
        return None, "CDE JSON not supplied"

    variable_key = normalize_identifier(selected_variable)
    cde_key = normalize_identifier(selected_cde_name)

    variable_matches = lookup.get(variable_key, []) if variable_key else []
    if len(variable_matches) == 1:
        return variable_matches[0], "Matched by final CDE variable"
    if len(variable_matches) > 1:
        narrowed = [
            entry
            for entry in variable_matches
            if normalize_identifier(entry.get("CDE Name")) == cde_key
        ]
        if len(narrowed) == 1:
            return narrowed[0], "Matched by final CDE variable and name"
        return variable_matches[0], "Ambiguous variable lookup; first official entry used"

    cde_matches = lookup.get(cde_key, []) if cde_key else []
    if len(cde_matches) == 1:
        return cde_matches[0], "Matched by final CDE name"
    if len(cde_matches) > 1:
        return cde_matches[0], "Ambiguous CDE-name lookup; first official entry used"
    return None, "Final CDE was not found in the supplied JSON"


def recompute_concept_score(
    study_variable: Any,
    study_description: Any,
    study_form: Any,
    expected_crf: Any,
    final_variable: Any,
    official_question: Any,
    final_crf: Any,
    settings: dict[str, float],
) -> float:
    score = (
        similarity_score(study_variable, final_variable) * settings["var_name_weight"]
        + similarity_score(study_description, official_question) * settings["field_label_weight"]
        + similarity_score(study_form, final_crf) * settings["form_context_weight"]
        + similarity_score(expected_crf, final_crf) * settings["crf_context_weight"]
    )
    if (
        normalize_crf_label(expected_crf)
        and normalize_crf_label(expected_crf) == normalize_crf_label(final_crf)
    ):
        score += settings["crf_concept_bonus"]
    return round(min(score, 100.0), 1)


def classify_combinability(
    decision: str,
    recon_confidence: str,
    concept_score: float | None,
    fidelity_score: float | None,
    encoding_scorable: bool,
    settings: dict[str, float],
) -> str:
    if decision == "Needs Human Review":
        return "Needs human review"
    if decision == "No Match" or not decision:
        return "No accepted CDE mapping"
    if decision != "Accept Candidate":
        return "Unrecognized reconciliation decision"
    if recon_confidence.lower() != "high":
        return "Accepted mapping, lower reconciliation confidence"
    if concept_score is None:
        return "Accepted mapping, concept score unavailable"
    if concept_score < settings["concept_high"]:
        return "Accepted mapping, concept review recommended"
    if not encoding_scorable:
        return "Concept aligned, encoding not scorable"
    if fidelity_score is not None and fidelity_score >= settings["fidelity_high"]:
        return "Ready to combine"
    if fidelity_score is not None and fidelity_score >= settings["fidelity_medium"]:
        return "Concept aligned, encoding transformation needed"
    return "Concept aligned, substantial encoding harmonization needed"


def valid_crf(value: Any) -> bool:
    text = clean_text(value)
    return bool(text) and text.lower() not in {
        "no crf match",
        "no confident heal core crf match exists",
        "none",
        "nan",
    }


def file_identifier(path: Path) -> str:
    name = re.sub(r"_cde_reconciled(?:\(\d+\))?$", "", path.stem, flags=re.I)
    return name


def extract_detail(
    workbook_path: Path,
    settings: dict[str, float],
    cde_lookup: dict[str, list[dict[str, Any]]],
) -> pd.DataFrame:
    try:
        frame = pd.read_excel(workbook_path, sheet_name="all-outputs")
    except ValueError as exc:
        raise ValueError(f"{workbook_path.name}: missing required 'all-outputs' sheet") from exc

    columns = list(frame.columns)
    resolved_variable_col = first_existing(columns, ["Resolved Variable Column"])
    resolved_description_col = first_existing(columns, ["Resolved Description Column"])
    resolved_form_col = first_existing(columns, ["Resolved CRF Column"])
    resolved_encoding_col = first_existing(columns, ["Resolved Encoding Column"])

    variable_col = (
        clean_text(frame[resolved_variable_col].dropna().iloc[0])
        if resolved_variable_col and not frame[resolved_variable_col].dropna().empty
        else first_existing(columns, ["name", "variable", "var"])
    )
    description_col = (
        clean_text(frame[resolved_description_col].dropna().iloc[0])
        if resolved_description_col and not frame[resolved_description_col].dropna().empty
        else first_existing(columns, ["description", "desc", "title"])
    )
    form_col = (
        clean_text(frame[resolved_form_col].dropna().iloc[0])
        if resolved_form_col and not frame[resolved_form_col].dropna().empty
        else first_existing(columns, ["section", "form", "module"])
    )
    encoding_col = (
        clean_text(frame[resolved_encoding_col].dropna().iloc[0])
        if resolved_encoding_col and not frame[resolved_encoding_col].dropna().empty
        else first_existing(columns, ["enumLabels", "constraints", "encoding", "encodings"])
    )

    for required in ["recon_decision", "best_best_match_variable", "best_best_match_cde", "best_best_match_crf"]:
        if required not in columns:
            raise KeyError(f"{workbook_path.name}: required column '{required}' was not found")

    records: list[dict[str, Any]] = []
    for source_index, row in frame.iterrows():
        decision = clean_text(row.get("recon_decision"))
        recon_confidence = clean_text(row.get("recon_confidence"))
        final_variable = clean_text(row.get("best_best_match_variable"))
        final_cde_name = clean_text(row.get("best_best_match_cde"))
        final_crf = clean_text(row.get("best_best_match_crf"))

        study_variable = clean_text(row.get(variable_col)) if variable_col else ""
        study_description = clean_text(row.get(description_col)) if description_col else ""
        study_form = clean_text(row.get(form_col)) if form_col else ""
        study_encoding = value_to_text(row.get(encoding_col)) if encoding_col else ""
        expected_crf = clean_text(row.get("HEAL Core CRF Match"))
        if not valid_crf(expected_crf):
            expected_crf = ""

        official_entry, lookup_status = choose_official_entry(
            cde_lookup, final_variable, final_cde_name
        )
        official_question = ""
        official_pv_description = ""
        if official_entry is not None:
            official_question = value_to_text(
                official_entry.get("Additional Notes (Question Text)")
                or official_entry.get("additional_notes_question_text")
            )
            official_pv_description = value_to_text(
                official_entry.get("PV Description")
                or official_entry.get("pv_description")
            )

        recomputed_concept = None
        recomputed_fidelity = None
        encoding_scorable = bool(study_encoding and official_pv_description)
        if official_entry is not None and final_variable:
            recomputed_concept = recompute_concept_score(
                study_variable,
                study_description,
                study_form,
                expected_crf,
                final_variable,
                official_question,
                final_crf,
                settings,
            )
            if encoding_scorable:
                recomputed_fidelity = float(
                    similarity_score(study_encoding, official_pv_description)
                )

        initial_candidate = clean_text(row.get("Final HEAL CDE Concept Match"))
        initial_aligned = bool(
            final_variable
            and initial_candidate
            and normalize_identifier(final_variable) == normalize_identifier(initial_candidate)
        )

        concept_for_tier = recomputed_concept
        fidelity_for_tier = recomputed_fidelity
        score_source = "Recomputed against final reconciled CDE" if official_entry else "Unavailable"
        if official_entry is None and initial_aligned:
            concept_for_tier = safe_float(row.get("Final Concept Match Score"))
            fidelity_for_tier = safe_float(row.get("Final Encoding Fidelity Score"))
            encoding_scorable = bool(study_encoding)
            score_source = "Existing score; final CDE equals retrieval candidate"

        tier = classify_combinability(
            decision,
            recon_confidence,
            concept_for_tier,
            fidelity_for_tier,
            encoding_scorable,
            settings,
        )

        records.append(
            {
                "file_id": file_identifier(workbook_path),
                "source_workbook": workbook_path.name,
                "source_row_number": int(source_index) + 2,
                "original_variable": study_variable,
                "original_form": study_form,
                "original_description": study_description,
                "original_encoding": study_encoding,
                "prestep_heal_core_crf": clean_text(row.get("HEAL Core CRF Match")),
                "prestep_crf_confidence": clean_text(row.get("Prestep CRF Confidence")),
                "initial_cde_candidate": initial_candidate,
                "existing_final_concept_score": safe_float(row.get("Final Concept Match Score")),
                "existing_final_encoding_fidelity_score": safe_float(row.get("Final Encoding Fidelity Score")),
                "final_heal_core_crf": final_crf,
                "final_cde_name": final_cde_name,
                "final_cde_variable": final_variable,
                "recon_decision": decision,
                "recon_confidence": recon_confidence,
                "recon_result_source": clean_text(row.get("recon_result_source")),
                "recon_rationale": clean_text(row.get("recon_rationale")),
                "initial_candidate_matches_final_variable": initial_aligned,
                "cde_json_lookup_status": lookup_status,
                "score_source": score_source,
                "recomputed_final_concept_similarity": recomputed_concept,
                "encoding_scorable": encoding_scorable,
                "recomputed_final_encoding_fidelity": recomputed_fidelity,
                "combinability_tier": tier,
                "counts_as_ready_to_combine": tier == "Ready to combine",
            }
        )
    return pd.DataFrame.from_records(records)


def safe_rate(numerator: int, denominator: int) -> float | None:
    return numerator / denominator if denominator else None


def mean_or_none(series: pd.Series) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    return round(float(numeric.mean()), 1) if not numeric.empty else None


def median_or_none(series: pd.Series) -> float | None:
    numeric = pd.to_numeric(series, errors="coerce").dropna()
    return round(float(numeric.median()), 1) if not numeric.empty else None


def summarize_file(detail: pd.DataFrame) -> dict[str, Any]:
    total = len(detail)
    accepted = detail["recon_decision"].eq("Accept Candidate") & detail["final_cde_variable"].ne("")
    accepted_high = accepted & detail["recon_confidence"].str.lower().eq("high")
    needs_review = detail["recon_decision"].eq("Needs Human Review")
    no_match = detail["recon_decision"].eq("No Match")
    initial_candidate = detail["initial_cde_candidate"].ne("")
    aligned = accepted & detail["initial_candidate_matches_final_variable"]
    recomputed = accepted_high & detail["recomputed_final_concept_similarity"].notna()
    encoding_scorable = recomputed & detail["encoding_scorable"]
    ready = detail["counts_as_ready_to_combine"]

    detected_crfs = sorted(
        {
            value
            for value in detail["prestep_heal_core_crf"].tolist()
            if valid_crf(value)
        }
    )
    accepted_crfs = sorted(
        {
            value
            for value in detail.loc[accepted, "final_heal_core_crf"].tolist()
            if valid_crf(value)
        }
    )

    return {
        "file_id": detail["file_id"].iloc[0],
        "source_workbook": detail["source_workbook"].iloc[0],
        "total_variables": total,
        "detected_heal_core_crf_count": len(detected_crfs),
        "detected_heal_core_crfs": " | ".join(detected_crfs),
        "accepted_mapping_crf_count": len(accepted_crfs),
        "accepted_mapping_crfs": " | ".join(accepted_crfs),
        "initial_cde_candidate_count": int(initial_candidate.sum()),
        "accepted_cde_mapping_count": int(accepted.sum()),
        "accepted_cde_mapping_rate": safe_rate(int(accepted.sum()), total),
        "high_confidence_accepted_count": int(accepted_high.sum()),
        "high_confidence_accepted_rate_total": safe_rate(int(accepted_high.sum()), total),
        "high_confidence_share_of_accepted": safe_rate(int(accepted_high.sum()), int(accepted.sum())),
        "needs_human_review_count": int(needs_review.sum()),
        "no_match_count": int(no_match.sum()),
        "accepted_score_aligned_count": int(aligned.sum()),
        "accepted_score_not_aligned_count": int((accepted & ~aligned).sum()),
        "recomputed_concept_score_count": int(recomputed.sum()),
        "avg_recomputed_concept_similarity": mean_or_none(
            detail.loc[recomputed, "recomputed_final_concept_similarity"]
        ),
        "median_recomputed_concept_similarity": median_or_none(
            detail.loc[recomputed, "recomputed_final_concept_similarity"]
        ),
        "encoding_scorable_high_confidence_count": int(encoding_scorable.sum()),
        "avg_recomputed_encoding_fidelity": mean_or_none(
            detail.loc[encoding_scorable, "recomputed_final_encoding_fidelity"]
        ),
        "median_recomputed_encoding_fidelity": median_or_none(
            detail.loc[encoding_scorable, "recomputed_final_encoding_fidelity"]
        ),
        "ready_to_combine_count": int(ready.sum()),
        "ready_to_combine_rate_scorable": safe_rate(int(ready.sum()), int(encoding_scorable.sum())),
        "ready_to_combine_rate_total_variables": safe_rate(int(ready.sum()), total),
    }


def output_basename(value: Any) -> str:
    text = clean_text(value)
    return re.split(r"[\\/]", text)[-1] if text else ""


def load_run_report(path: Path | None) -> pd.DataFrame | None:
    if path is None:
        return None
    if not path.exists():
        raise FileNotFoundError(f"Run report not found: {path}")
    frame = pd.read_csv(path)
    if "output_file" in frame.columns:
        frame["output_basename"] = frame["output_file"].map(output_basename)
    return frame


def add_run_status(summary: pd.DataFrame, run_report: pd.DataFrame | None) -> pd.DataFrame:
    result = summary.copy()
    if run_report is None or "output_basename" not in run_report.columns:
        result["run_status"] = ""
        result["run_field_count"] = None
        result["processing_seconds"] = None
        return result

    lookup = run_report.drop_duplicates("output_basename", keep="last").set_index("output_basename")
    result["run_status"] = result["source_workbook"].map(lookup.get("status", pd.Series(dtype=object)))
    result["run_field_count"] = result["source_workbook"].map(lookup.get("field_count", pd.Series(dtype=object)))
    result["processing_seconds"] = result["source_workbook"].map(
        lookup.get("processing_seconds", pd.Series(dtype=object))
    )
    return result


def definitions_frame(settings: dict[str, float], used_cde_json: bool) -> pd.DataFrame:
    rows = [
        ("accepted_cde_mapping_count", "Rows with recon_decision = Accept Candidate and a final CDE variable."),
        ("accepted_cde_mapping_rate", "Accepted CDE mappings divided by all variable rows in the data dictionary."),
        ("high_confidence_accepted_count", "Accepted mappings where recon_confidence = High. This is model confidence, not validated accuracy."),
        ("recomputed_final_concept_similarity", "0-100 fuzzy concept score recalculated against the final reconciled CDE variable and official question text."),
        ("recomputed_final_encoding_fidelity", "0-100 fuzzy similarity between the study encoding and the official CDE PV Description. Blank when either side lacks scorable encoding text."),
        ("ready_to_combine_count", f"Accepted + High recon confidence + concept similarity >= {settings['concept_high']:.0f} + encoding fidelity >= {settings['fidelity_high']:.0f}."),
        ("ready_to_combine_rate_scorable", "Ready-to-combine rows divided by accepted, high-confidence rows with both study and official encodings available."),
        ("accepted_score_not_aligned_count", "Accepted rows where the final reconciled CDE variable differs from the initial retrieval candidate. Existing notebook scores should not be averaged for these rows without recomputation."),
        ("Accuracy", "Not calculated. True accuracy requires human-reviewed reference labels: confirmed correct mappings divided by reviewed proposed mappings."),
        ("HSS availability", "Not inferred from CDE-ID outputs. Join a separate HSS availability field before deciding which rows belong on the data package board."),
        ("CDE JSON used", "Yes" if used_cde_json else "No. Post-reconciliation recomputation will be unavailable or provisional."),
    ]
    return pd.DataFrame(rows, columns=["metric", "definition"])


def style_workbook(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    thin_gray = Side(style="thin", color=LIGHT_GRAY)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_row >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = PatternFill("solid", fgColor=HEAL_PURPLE)
                cell.font = Font(color=WHITE, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(bottom=Side(style="medium", color=HEAL_MAGENTA))
            worksheet.row_dimensions[1].height = 32

        for column_cells in worksheet.iter_cols():
            letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells[: min(len(column_cells), 300)]:
                text = clean_text(cell.value)
                max_length = max(max_length, max((len(line) for line in text.splitlines()), default=0))
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                    cell.border = Border(bottom=thin_gray)
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 42)

        for row_number in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_number].height = 18

    if "file-summary" in workbook.sheetnames:
        sheet = workbook["file-summary"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        for header, column in headers.items():
            if header and ("rate" in str(header) or "share" in str(header)):
                for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
                    cell[0].number_format = "0.0%"
            elif header and ("avg_" in str(header) or "median_" in str(header)):
                for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
                    cell[0].number_format = "0.0"

        for metric in [
            "accepted_cde_mapping_rate",
            "high_confidence_share_of_accepted",
            "ready_to_combine_rate_scorable",
        ]:
            if metric in headers and sheet.max_row >= 2:
                letter = get_column_letter(headers[metric])
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    ColorScaleRule(
                        start_type="num", start_value=0, start_color=RED,
                        mid_type="num", mid_value=0.5, mid_color=YELLOW,
                        end_type="num", end_value=1, end_color=GREEN,
                    ),
                )

    if "variable-detail" in workbook.sheetnames:
        sheet = workbook["variable-detail"]
        headers = {cell.value: cell.column for cell in sheet[1]}
        for header in ["recomputed_final_concept_similarity", "recomputed_final_encoding_fidelity"]:
            if header in headers:
                column = headers[header]
                for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
                    cell[0].number_format = "0.0"
        tier_col = headers.get("combinability_tier")
        if tier_col:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=tier_col)
                value = clean_text(cell.value)
                if value == "Ready to combine":
                    cell.fill = PatternFill("solid", fgColor=GREEN)
                elif "review" in value.lower() or "lower" in value.lower():
                    cell.fill = PatternFill("solid", fgColor=YELLOW)
                elif "No accepted" in value:
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    if "metric-definitions" in workbook.sheetnames:
        sheet = workbook["metric-definitions"]
        sheet.column_dimensions["A"].width = 40
        sheet.column_dimensions["B"].width = 105
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
            sheet.row_dimensions[row].height = 36

    if "run-status" in workbook.sheetnames:
        sheet = workbook["run-status"]
        for row in range(2, sheet.max_row + 1):
            sheet.row_dimensions[row].height = 18

    styled_path = output_path.with_name(f"{output_path.stem}.styled{output_path.suffix}")
    workbook.save(styled_path)
    styled_path.replace(output_path)


def main() -> None:
    args = parse_args()
    if not args.input_dir.exists():
        raise FileNotFoundError(f"Input directory not found: {args.input_dir}")
    if args.output.suffix.lower() != ".xlsx":
        raise ValueError("--output must end in .xlsx")

    iterator = args.input_dir.rglob("*.xlsx") if args.recursive else args.input_dir.glob("*.xlsx")
    workbook_paths = sorted(
        path
        for path in iterator
        if re.search(r"_cde_reconciled(?:\(\d+\))?\.xlsx$", path.name, flags=re.I)
        and path.resolve() != args.output.resolve()
    )
    if not workbook_paths:
        raise FileNotFoundError(
            f"No *_cde_reconciled*.xlsx files were found in {args.input_dir}"
        )

    settings = read_settings(args.config)
    cde_lookup = load_cde_lookup(args.cde_json)
    run_report = load_run_report(args.run_report)

    detail_frames = []
    error_records = []
    for workbook_path in workbook_paths:
        print(f"Reading {workbook_path.name}...")
        try:
            detail_frames.append(extract_detail(workbook_path, settings, cde_lookup))
        except Exception as exc:
            error_records.append(
                {"source_workbook": workbook_path.name, "report_error": str(exc)}
            )

    if not detail_frames:
        raise RuntimeError("No workbooks could be summarized. Check the report-errors output.")

    detail = pd.concat(detail_frames, ignore_index=True)
    summary_audit = pd.DataFrame(
        [summarize_file(group) for _, group in detail.groupby("source_workbook", sort=True)]
    )
    summary_audit = add_run_status(summary_audit, run_report)
    summary_core_columns = [
        "file_id",
        "source_workbook",
        "run_status",
        "total_variables",
        "detected_heal_core_crf_count",
        "detected_heal_core_crfs",
        "accepted_mapping_crf_count",
        "accepted_mapping_crfs",
        "accepted_cde_mapping_count",
        "accepted_cde_mapping_rate",
        "high_confidence_accepted_count",
        "high_confidence_share_of_accepted",
        "needs_human_review_count",
        "no_match_count",
        "accepted_score_not_aligned_count",
        "avg_recomputed_concept_similarity",
        "encoding_scorable_high_confidence_count",
        "avg_recomputed_encoding_fidelity",
        "ready_to_combine_count",
        "ready_to_combine_rate_scorable",
    ]
    summary = summary_audit[summary_core_columns].copy()
    definitions = definitions_frame(settings, bool(cde_lookup))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="file-summary", index=False)
        summary_audit.to_excel(writer, sheet_name="file-summary-audit", index=False)
        detail.to_excel(writer, sheet_name="variable-detail", index=False)
        definitions.to_excel(writer, sheet_name="metric-definitions", index=False)
        if run_report is not None:
            run_report.to_excel(writer, sheet_name="run-status", index=False)
        if error_records:
            pd.DataFrame(error_records).to_excel(writer, sheet_name="report-errors", index=False)

    style_workbook(args.output)
    print(f"\nReport created: {args.output}")
    print(f"Workbooks summarized: {len(summary)}")
    print(f"Variable rows summarized: {len(detail)}")
    if error_records:
        print(f"Workbooks with report errors: {len(error_records)}")


if __name__ == "__main__":
    main()
